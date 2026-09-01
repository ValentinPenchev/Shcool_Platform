import io
import openpyxl

_CHART_TYPE_NAMES = (
    "BarChart", "PieChart", "LineChart", "AreaChart",
    "ScatterChart", "DoughnutChart", "RadarChart", "BubbleChart"
)


def _iter_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            yield cell


def _has_chart(ws):
    return len(ws._charts) > 0


def _has_recognized_chart_type(ws):
    return any(type(chart).__name__ in _CHART_TYPE_NAMES for chart in ws._charts)


def _has_bold(ws):
    return any(c.font and c.font.bold for c in _iter_cells(ws))


def _has_formulas(ws):
    return any(
        isinstance(c.value, str) and c.value.startswith("=")
        for c in _iter_cells(ws)
    )


def _has_absolute_reference(ws):
    for c in _iter_cells(ws):
        if isinstance(c.value, str) and c.value.startswith("=") and "$" in c.value:
            return True
    return False


def _has_custom_data_types(ws):
    return any(c.number_format and c.number_format != "General" for c in _iter_cells(ws))


def _has_merged_cells(ws):
    return len(ws.merged_cells.ranges) > 0


def _has_borders(ws):
    for c in _iter_cells(ws):
        border = c.border
        if not border:
            continue
        for side in (border.left, border.right, border.top, border.bottom):
            if side and side.style:
                return True
    return False


def _has_font_color(ws):
    for c in _iter_cells(ws):
        try:
            rgb = c.font.color.rgb if c.font and c.font.color else None
        except Exception:
            rgb = None
        if isinstance(rgb, str) and rgb not in ("00000000", "FF000000"):
            return True
    return False


def _has_cell_shading(ws):
    for c in _iter_cells(ws):
        fill = c.fill
        if not fill or not fill.patternType or fill.patternType == "none":
            continue
        fg = getattr(fill.fgColor, "rgb", None)
        if isinstance(fg, str) and fg not in ("00000000",):
            return True
    return False


def _has_sort_filter(ws):
    return bool(ws.auto_filter and ws.auto_filter.ref)


def _has_summary_report(ws):
    for dim in ws.row_dimensions.values():
        if dim.outlineLevel and dim.outlineLevel > 0:
            return True
    for c in _iter_cells(ws):
        if isinstance(c.value, str):
            formula = c.value.upper().lstrip("=")
            if formula.startswith(("SUBTOTAL", "SUMIF", "SUMIFS")):
                return True
    try:
        if ws.parent._pivots:
            return True
    except Exception:
        pass
    return False


def evaluate_excel(file_bytes: bytes, criteria: dict) -> dict:
    """Оценява .xlsx файл спрямо предоставените критерии."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    ws = wb.active

    score = 0
    max_score = 0
    details = []

    def add_check(key, label, passed_note, failed_note, passed):
        nonlocal score, max_score
        rule = criteria[key]
        rule_points = rule.get("points", 1) if isinstance(rule, dict) else 1
        max_score += rule_points
        if passed:
            score += rule_points
        details.append({
            "criterion": label,
            "passed": passed,
            "score": rule_points if passed else 0,
            "max": rule_points,
            "note": passed_note if passed else failed_note
        })

    if "formulas" in criteria:
        add_check("formulas", "Използване на формули",
                   "Намерени са валидни Excel формули.",
                   "Не са намерени формули в електронната таблица.", _has_formulas(ws))

    if "chart" in criteria:
        add_check("chart", "Наличие на диаграма",
                   "Намерена е диаграма в работния лист.",
                   "Липсва диаграма в работния лист.", _has_chart(ws))

    if "chart_type" in criteria:
        add_check("chart_type", "Вид на диаграмата",
                   "Диаграмата е с разпознат тип.",
                   "Не е открита диаграма с разпознат тип.", _has_recognized_chart_type(ws))

    if "min_columns" in criteria:
        rule = criteria["min_columns"]
        required = rule.get("count", 1) if isinstance(rule, dict) else 1
        add_check("min_columns", "Брой колони",
                   f"Таблицата съдържа {ws.max_column} колони (минимално изисквани: {required}).",
                   f"Таблицата съдържа само {ws.max_column} колони (минимално изисквани: {required}).",
                   ws.max_column >= required)

    if "min_rows" in criteria:
        rule = criteria["min_rows"]
        required = rule.get("count", 1) if isinstance(rule, dict) else 1
        add_check("min_rows", "Брой редове",
                   f"Таблицата съдържа {ws.max_row} реда (минимално изисквани: {required}).",
                   f"Таблицата съдържа само {ws.max_row} реда (минимално изисквани: {required}).",
                   ws.max_row >= required)

    if "bold" in criteria:
        add_check("bold", "Удебеляване",
                   "Намерен е удебелен текст в клетки.",
                   "Не е намерен удебелен текст в клетки.", _has_bold(ws))

    if "absolute_reference" in criteria:
        add_check("absolute_reference", "Абсолютен/относителен адрес",
                   "Намерена е формула с абсолютен адрес ($).",
                   "Не е намерена формула с абсолютен адрес.", _has_absolute_reference(ws))

    if "data_types" in criteria:
        add_check("data_types", "Типове данни",
                   "Намерен е зададен формат на клетки (напр. валута, число).",
                   "Не е намерен зададен формат на клетки.", _has_custom_data_types(ws))

    if "merged_cells" in criteria:
        add_check("merged_cells", "Обединени клетки",
                   "Намерени са обединени клетки.",
                   "Не са намерени обединени клетки.", _has_merged_cells(ws))

    if "borders" in criteria:
        add_check("borders", "Добавени рамки",
                   "Намерени са добавени рамки на клетки.",
                   "Не са намерени рамки на клетки.", _has_borders(ws))

    if "font_color" in criteria:
        add_check("font_color", "Цвят на текста",
                   "Намерен е зададен цвят на текста.",
                   "Не е намерен зададен цвят на текста.", _has_font_color(ws))

    if "cell_shading" in criteria:
        add_check("cell_shading", "Оцветяване на клетка",
                   "Намерена е оцветена клетка.",
                   "Не е намерена оцветена клетка.", _has_cell_shading(ws))

    if "sort_filter" in criteria:
        add_check("sort_filter", "Сортиране и филтриране",
                   "Намерен е приложен филтър върху данните.",
                   "Не е намерен приложен филтър върху данните.", _has_sort_filter(ws))

    if "summary_report" in criteria:
        add_check("summary_report", "Обобщена справка",
                   "Намерена е обобщена справка (групиране на редове или обобщаваща формула).",
                   "Не е намерена обобщена справка.", _has_summary_report(ws))

    percentage = round((score / max_score) * 100) if max_score > 0 else 0
    return {
        "score": score,
        "max_score": max_score,
        "percentage": percentage,
        "details": details
    }
