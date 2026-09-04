import io
import os
import json
import time
import hashlib
import uuid
import asyncio
from datetime import datetime
from typing import Optional
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import pandas as pd
import docx

from supabase_client import supabase, upload_to_supabase, cleanup_expired_files, sanitize_storage_segment, BUCKET_NAME
from evaluators.word_eval import evaluate_word
from evaluators.excel_eval import evaluate_excel
from evaluators.ppt_eval import evaluate_ppt
from evaluators.metadata import extract_office_metadata
from evaluators.criteria_parser import extract_criteria_from_text

app = FastAPI(title="Office & Code Evaluator API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Парола за админ панела - подадена от клиента в хедъра X-Admin-Password. Пази се
# като environment variable в Render (може да се смени без промяна на кода), с fallback
# към стойността по подразбиране. Пазят се само маршрутите под /api/admin/* - линковете
# за ученици (задачи/качвания) остават публични и без парола.
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "972700")

@app.middleware("http")
async def admin_auth_middleware(request: Request, call_next):
    if request.method != "OPTIONS" and request.url.path.startswith("/api/admin"):
        provided = request.headers.get("x-admin-password")
        if provided != ADMIN_PASSWORD:
            return JSONResponse(status_code=401, content={"detail": "Невалидна или липсваща парола за достъп."})
    return await call_next(request)

# Много прост in-memory rate limit за публичните endpoint-и за качване (без парола) -
# защита от случайно/автоматизирано спамене на заявки. Нулира се при рестарт на сървъра,
# което е приемливо за мащаба на приложението (един сървър, без нужда от Redis).
_upload_rate_limit: dict[str, list[float]] = {}
RATE_LIMIT_MAX_REQUESTS = 15
RATE_LIMIT_WINDOW_SECONDS = 300

def _check_rate_limit(request: Request):
    client_ip = request.client.host if request.client else "unknown"
    now = time.time()
    timestamps = [t for t in _upload_rate_limit.get(client_ip, []) if now - t < RATE_LIMIT_WINDOW_SECONDS]
    if len(timestamps) >= RATE_LIMIT_MAX_REQUESTS:
        raise HTTPException(status_code=429, detail="Твърде много заявки. Опитайте отново след няколко минути.")
    timestamps.append(now)
    _upload_rate_limit[client_ip] = timestamps

# Материалите се пазят 60 дни, след което се изтриват автоматично (файл + запис)
CLEANUP_AFTER_DAYS = 60
CLEANUP_INTERVAL_SECONDS = 24 * 60 * 60

# Стандартна българска скала за оценяване (2-6), изведена от процента успех.
# Оценката не се пази отделно в базата - винаги се извежда от вече записания
# процент, за да не се налага промяна на схемата на submissions.
def _percentage_to_grade(percentage: float) -> dict:
    if percentage >= 91:
        grade, label = 6, "Отличен"
    elif percentage >= 76:
        grade, label = 5, "Много добър"
    elif percentage >= 61:
        grade, label = 4, "Добър"
    elif percentage >= 41:
        grade, label = 3, "Среден"
    else:
        grade, label = 2, "Слаб"
    return {"grade": grade, "grade_label": label}

def _parse_iso(ts):
    """Разбира ISO-8601 timestamp низ (както го връща Supabase), иначе None."""
    if not ts:
        return None
    try:
        return datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
    except Exception:
        return None

async def periodic_cleanup():
    while True:
        try:
            cleanup_expired_files(CLEANUP_AFTER_DAYS)
        except Exception as e:
            print(f"Забележка при почистване: {e}")
        await asyncio.sleep(CLEANUP_INTERVAL_SECONDS)

@app.on_event("startup")
async def startup_event():
    asyncio.create_task(periodic_cleanup())

@app.get("/api/health")
async def health_check():
    return {"status": "ok", "message": "Backend is running"}

# -----------------------------------------------------------------------------
# 1. УПРАВЛЕНИЕ НА КЛАСОВЕ (GROUPS - CRUD)
# -----------------------------------------------------------------------------

@app.get("/api/admin/groups")
async def get_all_groups():
    try:
        res = supabase.table("groups").select("*").execute()
        return res.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при четене на класовете: {str(e)}")

@app.post("/api/admin/groups")
async def create_or_update_group(
    group_id: str = Form(...),
    group_name: str = Form(...),
    students_json: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    inactive_students_json: Optional[str] = Form(None),
    avatars_json: Optional[str] = Form(None)
):
    students = []
    if students_json and students_json.strip():
        try:
            parsed = json.loads(students_json)
            if isinstance(parsed, list):
                students = [str(item).strip() for item in parsed if str(item).strip()]
        except Exception:
            students = [line.strip() for line in students_json.split("\n") if line.strip()]

    if not students and file:
        contents = await file.read()
        filename = file.filename.lower()
        try:
            if filename.endswith(".csv"):
                df = pd.read_csv(io.BytesIO(contents))
            elif filename.endswith(".xlsx"):
                df = pd.read_excel(io.BytesIO(contents))
            else:
                raise HTTPException(status_code=400, detail="Поддържат се само .csv и .xlsx файлове.")

            students = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Грешка при четене на файла: {str(e)}")

    data = {
        "group_id": group_id,
        "group_name": group_name,
        "students_json": students
    }

    # inactive_students_json се изпраща само от груповите действия (деактивиране/
    # активиране) - при обикновен запис на класа не се подава и не се пипа в базата
    if inactive_students_json is not None:
        try:
            parsed_inactive = json.loads(inactive_students_json)
            if isinstance(parsed_inactive, list):
                data["inactive_students_json"] = [str(item).strip() for item in parsed_inactive if str(item).strip()]
        except Exception:
            pass

    # avatars_json се изпраща само при ръчна смяна на аватар (момче/момиче) за ученик -
    # мапинг име -> "boy"/"girl", не се пипа при обикновен запис на класа
    if avatars_json is not None:
        try:
            parsed_avatars = json.loads(avatars_json)
            if isinstance(parsed_avatars, dict):
                data["student_avatars_json"] = {
                    str(k): str(v) for k, v in parsed_avatars.items() if str(v) in ("boy", "girl")
                }
        except Exception:
            pass

    try:
        res = supabase.table("groups").upsert(data).execute()
        return {"status": "success", "data": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при запис в базата: {str(e)}")

@app.delete("/api/admin/groups/{group_id}")
async def delete_group(group_id: str):
    try:
        supabase.table("groups").delete().eq("group_id", group_id).execute()
        return {"status": "success", "message": f"Класът {group_id} е изтрит."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при изтриване на класа: {str(e)}")

# -----------------------------------------------------------------------------
# 2. УПРАВЛЕНИЕ НА ЗАДАЧИ (ASSIGNMENTS - CRUD & FILE CRITERIA)
# -----------------------------------------------------------------------------

@app.get("/api/admin/assignments")
async def get_all_assignments(group_id: Optional[str] = None):
    try:
        query = supabase.table("assignments").select("*")
        if group_id:
            query = query.eq("group_id", group_id)
        res = query.execute()
        assignments = res.data or []

        # Формиране на дигиталната структура за всеки запис
        for a in assignments:
            task_id = a.get("id", "")
            a["link"] = f"/index.html?id={task_id}" if task_id else ""

        return assignments
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при четене на задачите: {str(e)}")

@app.post("/api/admin/assignments")
async def create_assignment(
    title: str = Form(...),
    group_id: str = Form(...),
    criteria_json: Optional[str] = Form(None),
    criteria_file: Optional[UploadFile] = File(None),
    template_id: Optional[str] = Form(None),
    deadline: Optional[str] = Form(None),
    reference_link: Optional[str] = Form(None),
    reference_file: Optional[UploadFile] = File(None),
    assignment_id: Optional[str] = Form(None)
):
    """
    Създава нова задача, а ако е подаден assignment_id на съществуваща - я обновява
    (upsert със същото id), запазвайки досегашните критерии/материали, ако за тях
    не е подадено нещо ново - редакцията не трябва тихо да ги изтрива.
    """
    is_edit = bool(assignment_id)
    final_id = assignment_id or str(uuid.uuid4())[:8]

    existing_assignment = None
    if is_edit:
        existing_res = supabase.table("assignments").select("*").eq("id", final_id).execute()
        if not existing_res.data:
            raise HTTPException(status_code=404, detail="Задачата за редакция не е намерена.")
        existing_assignment = existing_res.data[0]

    criteria_parsed = {}

    # Шаблон с готови критерии - взима им предимство пред качен Word/ръчен JSON,
    # за да не се въвеждат наново едни и същи критерии за всяка паралелка
    if template_id:
        tmpl_res = supabase.table("assignment_templates").select("criteria_json").eq("id", template_id).execute()
        if tmpl_res.data:
            criteria_parsed = tmpl_res.data[0].get("criteria_json") or {}
    elif criteria_file and criteria_file.filename.lower().endswith(".docx"):
        try:
            content = await criteria_file.read()
            doc = docx.Document(io.BytesIO(content))
            extracted_texts = [p.text.strip() for p in doc.paragraphs if p.text.strip()]

            # Абзаците се разпознават по ключови думи (напр. "таблица", "изображение",
            # "формула") и се превръщат в ключовете, които evaluator-ите реално четат -
            # генеричните "criterion_N" по-рано не участваха в оценяването изобщо.
            criteria_parsed = extract_criteria_from_text(extracted_texts)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Грешка при четене на Word файла: {str(e)}")
    elif criteria_json and criteria_json.strip() and criteria_json.strip() != "{}":
        try:
            criteria_parsed = json.loads(criteria_json)
        except Exception:
            criteria_parsed = {}
    elif existing_assignment:
        # Редакция без нов файл/шаблон - критериите остават непроменени
        criteria_parsed = existing_assignment.get("criteria_json") or {}

    reference_file_url = existing_assignment.get("reference_file_url") if existing_assignment else None
    if reference_file and reference_file.filename:
        try:
            ref_contents = await reference_file.read()
            ref_path = f"materials/{sanitize_storage_segment(final_id)}/{sanitize_storage_segment(reference_file.filename)}"
            reference_file_url = upload_to_supabase(ref_contents, reference_file.filename, ref_path)
        except Exception as e:
            print(f"Грешка при качване на помощния материал: {e}")

    reference_link_value = reference_link.strip() if reference_link and reference_link.strip() else None
    if not reference_link_value and existing_assignment:
        reference_link_value = existing_assignment.get("reference_link")

    data = {
        "id": final_id,
        "title": title,
        "group_id": group_id,
        "criteria_json": criteria_parsed,
        "deadline": deadline or None,
        "reference_link": reference_link_value,
        "reference_file_url": reference_file_url
    }

    try:
        supabase.table("assignments").upsert(data).execute()
        return {
            "status": "success",
            "assignment_id": final_id,
            "title": title,
            "group_id": group_id,
            "link": f"/index.html?id={final_id}"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при {'обновяване' if is_edit else 'създаване'} на задача: {str(e)}")

@app.delete("/api/admin/assignments/{assignment_id}")
async def delete_assignment(assignment_id: str):
    try:
        supabase.table("assignments").delete().eq("id", assignment_id).execute()
        return {"status": "success", "message": f"Задачата {assignment_id} е изтрита."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при изтриване на задачата: {str(e)}")

# Шаблони за задачи (запазени критерии за повторна употреба между паралелки)
@app.get("/api/admin/templates")
async def get_templates():
    try:
        res = supabase.table("assignment_templates").select("*").order("created_at", desc=True).execute()
        return res.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при четене на шаблоните: {str(e)}")

@app.post("/api/admin/templates")
async def create_template(
    title: str = Form(...),
    criteria_json: str = Form(...)
):
    try:
        parsed = json.loads(criteria_json)
    except Exception:
        raise HTTPException(status_code=400, detail="Невалидни критерии за шаблон.")

    try:
        supabase.table("assignment_templates").insert({"title": title, "criteria_json": parsed}).execute()
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при запис на шаблона: {str(e)}")

@app.delete("/api/admin/templates/{template_id}")
async def delete_template(template_id: int):
    try:
        supabase.table("assignment_templates").delete().eq("id", template_id).execute()
        return {"status": "success", "message": f"Шаблонът {template_id} е изтрит."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при изтриване на шаблона: {str(e)}")

@app.get("/api/assignments/{assignment_id}")
async def get_assignment(assignment_id: str):
    res = supabase.table("assignments").select("*").eq("id", assignment_id).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Задачата не е намерена.")
    
    assignment = res.data[0]
    group_res = supabase.table("groups").select("*").eq("group_id", assignment["group_id"]).execute()
    
    students = group_res.data[0]["students_json"] if group_res.data and "students_json" in group_res.data[0] else []
    group_name = group_res.data[0]["group_name"] if group_res.data and "group_name" in group_res.data[0] else assignment["group_id"]

    return {
        "id": assignment["id"],
        "title": assignment["title"],
        "group_id": assignment["group_id"],
        "group_name": group_name,
        "students": students,
        "criteria": assignment.get("criteria_json", {}),
        "deadline": assignment.get("deadline"),
        "reference_file_url": assignment.get("reference_file_url"),
        "reference_link": assignment.get("reference_link")
    }

# -----------------------------------------------------------------------------
# 3. ОЦЕНЯВАНЕ И ТАБЛО (SUBMISSIONS)
# -----------------------------------------------------------------------------

@app.post("/api/evaluate")
async def evaluate_file(
    request: Request,
    class_id: str = Form(...),
    student_name: str = Form(...),
    criteria_json: str = Form(...),
    assignment_id: Optional[str] = Form(None),
    file: UploadFile = File(...)
):
    _check_rate_limit(request)
    contents = await file.read()
    file_hash = hashlib.md5(contents).hexdigest()
    meta = extract_office_metadata(contents)
    creation_time = meta.get("created")

    storage_path = f"{sanitize_storage_segment(class_id)}/{sanitize_storage_segment(student_name)}_{sanitize_storage_segment(file.filename)}"

    try:
        file_url = upload_to_supabase(contents, file.filename, storage_path)
    except Exception:
        file_url = "#"

    filename = file.filename.lower()
    try:
        criteria = json.loads(criteria_json)
    except Exception:
        criteria = {}
    
    if filename.endswith(".docx"):
        result = evaluate_word(contents, criteria)
    elif filename.endswith(".xlsx"):
        result = evaluate_excel(contents, criteria)
    elif filename.endswith(".pptx"):
        result = evaluate_ppt(contents, criteria)
    else:
        raise HTTPException(status_code=400, detail="Форматът не се поддържа.")

    percentage = round((result["score"] / result["max_score"] * 100), 1) if result["max_score"] > 0 else 0
    grade_info = _percentage_to_grade(percentage)

    # Забележка: таблицата submissions няма собствена колона assignment_id, затова
    # го пазим вътре в details_json (вече съществуваща JSON колона), за да не се
    # налага промяна на схемата на базата данни.
    submission_data = {
        "student_name": student_name,
        "class_id": class_id,
        "filename": file.filename,
        "file_url": file_url,
        "storage_path": storage_path,
        "file_hash": file_hash,
        "creation_time": str(creation_time) if creation_time else None,
        "score": result["score"],
        "max_score": result["max_score"],
        "percentage": percentage,
        "details_json": {"assignment_id": assignment_id, "details": result["details"]}
    }

    # Ако ученикът вече е предавал по тази задача, старото предаване се заменя (файл +
    # запис) вместо да се трупа като отделен ред - последното предадено решение важи.
    if assignment_id:
        try:
            existing_res = supabase.table("submissions").select("id, storage_path, details_json") \
                .eq("class_id", class_id).eq("student_name", student_name).execute()
            for row in (existing_res.data or []):
                row_details = row.get("details_json")
                row_assignment_id = row_details.get("assignment_id") if isinstance(row_details, dict) else None
                if row_assignment_id == assignment_id:
                    if row.get("storage_path"):
                        try:
                            supabase.storage.from_(BUCKET_NAME).remove([row["storage_path"]])
                        except Exception as e:
                            print(f"Забележка при изтриване на старо предаване от Storage: {e}")
                    supabase.table("submissions").delete().eq("id", row["id"]).execute()
        except Exception as e:
            print(f"Забележка при проверка за предишно предаване: {e}")

    try:
        supabase.table("submissions").insert(submission_data).execute()
    except Exception as e:
        print(f"Грешка при запис на предаването: {e}")

    return {
        "student_name": student_name,
        "class_id": class_id,
        "filename": file.filename,
        "file_url": file_url,
        "score": result["score"],
        "max_score": result["max_score"],
        "percentage": percentage,
        "grade": grade_info["grade"],
        "grade_label": grade_info["grade_label"],
        "details": result["details"]
    }

@app.get("/api/admin/submissions")
async def get_submissions(group_id: Optional[str] = None, assignment_id: Optional[str] = None):
    try:
        query = supabase.table("submissions").select("*")
        if group_id:
            query = query.eq("class_id", group_id)
        res = query.execute()
        rows = res.data or []

        # Крайни срокове по assignment_id - за автоматично маркиране на закъснели предавания
        deadlines_res = supabase.table("assignments").select("id, deadline").execute()
        deadline_by_assignment = {
            a["id"]: a.get("deadline") for a in (deadlines_res.data or []) if a.get("deadline")
        }

        # assignment_id се пази вътре в details_json - извличаме го тук за удобство
        for row in rows:
            details = row.get("details_json")
            row["assignment_id"] = details.get("assignment_id") if isinstance(details, dict) else None
            grade_info = _percentage_to_grade(row.get("percentage") or 0)
            row["grade"] = grade_info["grade"]
            row["grade_label"] = grade_info["grade_label"]

            deadline_dt = _parse_iso(deadline_by_assignment.get(row["assignment_id"]))
            submitted_dt = _parse_iso(row.get("created_at"))
            row["is_late"] = bool(deadline_dt and submitted_dt and submitted_dt > deadline_dt)

        if assignment_id:
            rows = [r for r in rows if r.get("assignment_id") == assignment_id]

        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при четене: {str(e)}")

@app.get("/api/admin/submissions/export")
async def export_submissions(
    group_id: Optional[str] = None,
    assignment_id: Optional[str] = None,
    search: Optional[str] = None
):
    """Изнася предадените решения като форматиран Excel файл (лист с резултати + обобщение)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    rows = await get_submissions(group_id=group_id, assignment_id=assignment_id)

    if search:
        needle = search.strip().lower()
        rows = [r for r in rows if needle in (r.get("student_name") or "").lower()]

    # Заглавия на задачите и имена на класовете, за да не се показват голи идентификатори
    try:
        assignments_res = supabase.table("assignments").select("id, title").execute()
        title_by_assignment = {a["id"]: a.get("title") for a in (assignments_res.data or [])}
    except Exception:
        title_by_assignment = {}
    try:
        groups_res = supabase.table("groups").select("group_id, group_name").execute()
        name_by_group = {g["group_id"]: g.get("group_name") for g in (groups_res.data or [])}
    except Exception:
        name_by_group = {}

    rows.sort(key=lambda r: (r.get("class_id") or "", r.get("student_name") or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Резултати"

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1E293B")
    muted_font = Font(size=10, color="64748B")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["№", "Ученик", "Клас", "Задача", "Файл", "Предадено на",
               "Точки", "Максимум", "Успех (%)", "Оценка", "Статус"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="Резултати от предадените задачи").font = title_font

    scope_bits = []
    if group_id:
        scope_bits.append(f"Клас: {name_by_group.get(group_id, group_id)}")
    if assignment_id:
        scope_bits.append(f"Задача: {title_by_assignment.get(assignment_id, assignment_id)}")
    scope_bits.append(f"Генериран на {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=" · ".join(scope_bits)).font = muted_font

    header_row = 4
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 26

    grade_fills = {
        6: PatternFill("solid", fgColor="DCFCE7"),
        5: PatternFill("solid", fgColor="E0F2FE"),
        4: PatternFill("solid", fgColor="FEF3C7"),
        3: PatternFill("solid", fgColor="FFEDD5"),
        2: PatternFill("solid", fgColor="FEE2E2"),
    }

    for i, sub in enumerate(rows, start=1):
        r = header_row + i
        submitted = _parse_iso(sub.get("created_at"))
        values = [
            i,
            sub.get("student_name") or "",
            name_by_group.get(sub.get("class_id"), sub.get("class_id") or ""),
            title_by_assignment.get(sub.get("assignment_id"), sub.get("assignment_id") or "—"),
            sub.get("filename") or "",
            submitted.strftime("%d.%m.%Y %H:%M") if submitted else "",
            sub.get("score") or 0,
            sub.get("max_score") or 0,
            round(sub.get("percentage") or 0, 1),
            f'{sub.get("grade", "")} ({sub.get("grade_label", "")})'.strip(),
            "Закъснял" if sub.get("is_late") else "В срок",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border = border
            if col in (1, 7, 8, 9):
                cell.alignment = Alignment(horizontal="center")
            elif col in (10, 11):
                cell.alignment = Alignment(horizontal="center")
        grade_fill = grade_fills.get(sub.get("grade"))
        if grade_fill:
            ws.cell(row=r, column=10).fill = grade_fill
        if sub.get("is_late"):
            ws.cell(row=r, column=11).font = Font(color="B91C1C", bold=True)

    widths = [5, 26, 14, 26, 30, 18, 9, 11, 11, 18, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"

    # Втори лист с обобщение - общо число, среден успех и разпределение на оценките
    ws2 = wb.create_sheet("Обобщение")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 16
    ws2.cell(row=1, column=1, value="Обобщение").font = title_font

    total = len(rows)
    avg = round(sum((s.get("percentage") or 0) for s in rows) / total, 1) if total else 0
    late = sum(1 for s in rows if s.get("is_late"))
    summary = [
        ("Общо предадени работи", total),
        ("Среден успех (%)", avg),
        ("Закъснели предавания", late),
        ("Различни ученици", len({s.get("student_name") for s in rows if s.get("student_name")})),
    ]
    for i, (label, value) in enumerate(summary, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value)

    ws2.cell(row=8, column=1, value="Разпределение на оценките").font = Font(bold=True, size=12)
    dist_header = 9
    for col, name in enumerate(["Оценка", "Брой"], start=1):
        cell = ws2.cell(row=dist_header, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for i, grade in enumerate([6, 5, 4, 3, 2], start=1):
        count = sum(1 for s in rows if s.get("grade") == grade)
        ws2.cell(row=dist_header + i, column=1, value=grade).alignment = Alignment(horizontal="center")
        ws2.cell(row=dist_header + i, column=2, value=count).alignment = Alignment(horizontal="center")
        fill = grade_fills.get(grade)
        if fill:
            ws2.cell(row=dist_header + i, column=1).fill = fill

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"rezultati_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.delete("/api/admin/submissions/{submission_id}")
async def delete_submission(submission_id: int):
    try:
        res = supabase.table("submissions").select("storage_path").eq("id", submission_id).execute()
        if res.data and res.data[0].get("storage_path"):
            try:
                supabase.storage.from_(BUCKET_NAME).remove([res.data[0]["storage_path"]])
            except Exception as e:
                print(f"Забележка при изтриване на файла от Storage: {e}")

        supabase.table("submissions").delete().eq("id", submission_id).execute()
        return {"status": "success", "message": f"Предаването {submission_id} е изтрито."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при изтриване: {str(e)}")

# -----------------------------------------------------------------------------
# 3. УПРАЖНЕНИЯ (свободни качвания за практика, без автоматична проверка)
# -----------------------------------------------------------------------------
# Отделени от submissions умишлено - тук не се оценяват критерии, а само се
# брои колко пъти ученикът е качил работа в класа. При 5 качвания получава
# автоматично оценка "Отличен" (6), изчислена винаги от текущия брой записи,
# а не пазена отделно, за да няма разминаване при изтриване на качване.
EXCELLENT_UPLOAD_THRESHOLD = 5
EXCELLENT_GRADE = 6
EXCELLENT_GRADE_LABEL = "Отличен"

@app.get("/api/groups/{group_id}")
async def get_group_public(group_id: str):
    res = supabase.table("groups").select("*").eq("group_id", group_id).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Класът не е намерен.")
    group = res.data[0]
    return {
        "group_id": group["group_id"],
        "group_name": group.get("group_name", group_id),
        "students": group.get("students_json", [])
    }

@app.post("/api/exercise/upload")
async def upload_exercise(
    request: Request,
    class_id: str = Form(...),
    student_name: str = Form(...),
    file: UploadFile = File(...)
):
    _check_rate_limit(request)
    contents = await file.read()
    storage_path = (
        f"exercises/{sanitize_storage_segment(class_id)}/"
        f"{sanitize_storage_segment(student_name)}_{uuid.uuid4().hex[:6]}_{sanitize_storage_segment(file.filename)}"
    )

    try:
        file_url = upload_to_supabase(contents, file.filename, storage_path)
    except Exception:
        file_url = "#"

    try:
        supabase.table("exercise_uploads").insert({
            "class_id": class_id,
            "student_name": student_name,
            "filename": file.filename,
            "storage_path": storage_path,
            "file_url": file_url,
        }).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при запис на качването: {str(e)}")

    count_res = supabase.table("exercise_uploads").select("id") \
        .eq("class_id", class_id).eq("student_name", student_name).execute()
    count = len(count_res.data or [])

    return {
        "status": "success",
        "filename": file.filename,
        "count": count,
        "remaining": max(0, EXCELLENT_UPLOAD_THRESHOLD - count),
        "excellent": count >= EXCELLENT_UPLOAD_THRESHOLD
    }

@app.get("/api/admin/exercises")
async def get_exercise_uploads(group_id: Optional[str] = None):
    try:
        query = supabase.table("exercise_uploads").select("*")
        if group_id:
            query = query.eq("class_id", group_id)
        res = query.order("created_at", desc=True).execute()
        return res.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при четене на упражненията: {str(e)}")

@app.delete("/api/admin/exercises/{upload_id}")
async def delete_exercise_upload(upload_id: int):
    try:
        res = supabase.table("exercise_uploads").select("storage_path").eq("id", upload_id).execute()
        if res.data and res.data[0].get("storage_path"):
            try:
                supabase.storage.from_(BUCKET_NAME).remove([res.data[0]["storage_path"]])
            except Exception as e:
                print(f"Забележка при изтриване на файла от Storage: {e}")

        supabase.table("exercise_uploads").delete().eq("id", upload_id).execute()
        return {"status": "success", "message": f"Качването {upload_id} е изтрито."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при изтриване: {str(e)}")

@app.get("/api/admin/exercise-grades")
async def get_exercise_grades(group_id: Optional[str] = None):
    try:
        query = supabase.table("exercise_grade_log").select("*")
        if group_id:
            query = query.eq("class_id", group_id)
        res = query.order("entered_at", desc=True).execute()
        return res.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при четене на въведените оценки: {str(e)}")

@app.post("/api/admin/exercises/mark-graded")
async def mark_exercise_batch_graded(
    class_id: str = Form(...),
    student_name: str = Form(...)
):
    """
    Взима най-старите 5 (все още неотбелязани) качвания на ученика, трайно записва
    оценката в exercise_grade_log (с имената на файловете за архив), след което
    трие тези 5 качвания от exercise_uploads (файл + запис), за да не се трупат
    безкрайно в базата и Storage.
    """
    res = supabase.table("exercise_uploads").select("*") \
        .eq("class_id", class_id).eq("student_name", student_name) \
        .order("created_at", desc=False).execute()
    uploads = res.data or []

    if len(uploads) < EXCELLENT_UPLOAD_THRESHOLD:
        raise HTTPException(
            status_code=400,
            detail=f"Няма достатъчно качвания за въвеждане на оценка ({len(uploads)}/{EXCELLENT_UPLOAD_THRESHOLD})."
        )

    batch = uploads[:EXCELLENT_UPLOAD_THRESHOLD]
    filenames = [u.get("filename") for u in batch]
    ids_to_delete = [u["id"] for u in batch]
    paths_to_delete = [u["storage_path"] for u in batch if u.get("storage_path")]

    try:
        supabase.table("exercise_grade_log").insert({
            "class_id": class_id,
            "student_name": student_name,
            "grade": EXCELLENT_GRADE,
            "filenames": filenames
        }).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при записване на оценката: {str(e)}")

    if paths_to_delete:
        try:
            supabase.storage.from_(BUCKET_NAME).remove(paths_to_delete)
        except Exception as e:
            print(f"Забележка при изтриване на файловете от Storage: {e}")

    supabase.table("exercise_uploads").delete().in_("id", ids_to_delete).execute()

    remaining_res = supabase.table("exercise_uploads").select("id") \
        .eq("class_id", class_id).eq("student_name", student_name).execute()

    return {
        "status": "success",
        "grade": EXCELLENT_GRADE,
        "grade_label": EXCELLENT_GRADE_LABEL,
        "remaining_uploads": len(remaining_res.data or [])
    }

# -----------------------------------------------------------------------------
# 4. ПРИСЪСТВИЕ (по клас и дата)
# -----------------------------------------------------------------------------

@app.get("/api/admin/attendance")
async def get_attendance(group_id: str, record_date: str):
    try:
        res = supabase.table("attendance_records").select("*") \
            .eq("class_id", group_id).eq("record_date", record_date).execute()
        return res.data or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при четене на присъствието: {str(e)}")

@app.post("/api/admin/attendance")
async def mark_attendance(
    class_id: str = Form(...),
    student_name: str = Form(...),
    record_date: str = Form(...),
    status: str = Form(...)
):
    if status not in ("present", "absent"):
        raise HTTPException(status_code=400, detail="Невалиден статус.")
    try:
        supabase.table("attendance_records").upsert({
            "class_id": class_id,
            "student_name": student_name,
            "record_date": record_date,
            "status": status
        }, on_conflict="class_id,student_name,record_date").execute()
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при отбелязване на присъствие: {str(e)}")

@app.post("/api/admin/attendance/bulk")
async def mark_attendance_bulk(
    class_id: str = Form(...),
    record_date: str = Form(...),
    status: str = Form(...),
    students_json: str = Form(...)
):
    if status not in ("present", "absent"):
        raise HTTPException(status_code=400, detail="Невалиден статус.")
    try:
        students = json.loads(students_json)
        if not isinstance(students, list):
            raise ValueError
    except Exception:
        raise HTTPException(status_code=400, detail="Невалиден списък с ученици.")

    try:
        rows = [
            {"class_id": class_id, "student_name": name, "record_date": record_date, "status": status}
            for name in students
        ]
        if rows:
            supabase.table("attendance_records").upsert(rows, on_conflict="class_id,student_name,record_date").execute()
        return {"status": "success", "count": len(rows)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при масово отбелязване на присъствие: {str(e)}")

# -----------------------------------------------------------------------------
# 4.1. КАЛЕНДАР СЪС СЪБИТИЯ (собствени събития на учителя - родителски срещи,
#      контролни, ваканции). Сроковете на задачите се добавят от самия интерфейс,
#      затова тук се пазят само ръчно въведените събития.
# -----------------------------------------------------------------------------

@app.get("/api/admin/calendar-events")
async def get_calendar_events():
    # Таблицата calendar_events може още да не е създадена в Supabase - в такъв случай
    # календарът просто показва само сроковете на задачите, вместо да връща грешка.
    try:
        res = supabase.table("calendar_events").select("*").order("event_date").execute()
        return res.data or []
    except Exception as e:
        print(f"Забележка при четене на събитията от календара: {e}")
        return []

@app.post("/api/admin/calendar-events")
async def create_calendar_event(
    title: str = Form(...),
    event_date: str = Form(...),
    event_type: Optional[str] = Form("event"),
    class_id: Optional[str] = Form(None)
):
    if event_type not in ("event", "exam", "meeting", "holiday"):
        event_type = "event"
    try:
        res = supabase.table("calendar_events").insert({
            "title": title,
            "event_date": event_date,
            "event_type": event_type,
            "class_id": class_id or None
        }).execute()
        return {"status": "success", "data": res.data}
    except Exception as e:
        message = str(e)
        # Докато таблицата не е създадена в Supabase, календарът работи само със
        # сроковете на задачите - тук казваме ясно какво липсва, вместо суровата грешка
        if "calendar_events" in message and ("PGRST205" in message or "Could not find the table" in message):
            raise HTTPException(
                status_code=503,
                detail="Таблицата calendar_events още не е създадена в Supabase. Изпълнете SQL заявката за нея и опитайте отново."
            )
        raise HTTPException(status_code=500, detail=f"Грешка при запис на събитието: {message}")

@app.delete("/api/admin/calendar-events/{event_id}")
async def delete_calendar_event(event_id: int):
    try:
        supabase.table("calendar_events").delete().eq("id", event_id).execute()
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при изтриване на събитието: {str(e)}")

# -----------------------------------------------------------------------------
# 5. ЕМОЦИОМЕТЪР (дневно гласуване по клас - без нужда от парола, за да могат
#    учениците да гласуват направо от споделен линк, само нулирането е за учителя)
# -----------------------------------------------------------------------------

EMOTION_LIST = ["Щастлив", "Тъжен", "Кисел", "Доволен", "Любопитен", "Притеснен", "Влюбен"]

@app.get("/api/emotions")
async def get_emotions(group_id: str, record_date: str):
    try:
        res = supabase.table("emotion_votes").select("*") \
            .eq("class_id", group_id).eq("record_date", record_date).execute()
        counts = {row["emotion"]: row["count"] for row in (res.data or [])}
        return {emotion: counts.get(emotion, 0) for emotion in EMOTION_LIST}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при четене на емоциите: {str(e)}")

@app.post("/api/emotions/vote")
async def vote_emotion(
    request: Request,
    class_id: str = Form(...),
    record_date: str = Form(...),
    emotion: str = Form(...)
):
    _check_rate_limit(request)
    if emotion not in EMOTION_LIST:
        raise HTTPException(status_code=400, detail="Невалидна емоция.")
    try:
        existing = supabase.table("emotion_votes").select("id, count") \
            .eq("class_id", class_id).eq("record_date", record_date).eq("emotion", emotion).execute()
        if existing.data:
            new_count = existing.data[0]["count"] + 1
            supabase.table("emotion_votes").update({"count": new_count}).eq("id", existing.data[0]["id"]).execute()
        else:
            new_count = 1
            supabase.table("emotion_votes").insert({
                "class_id": class_id, "record_date": record_date, "emotion": emotion, "count": 1
            }).execute()
        return {"status": "success", "emotion": emotion, "count": new_count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при гласуване: {str(e)}")

@app.post("/api/admin/emotions/reset")
async def reset_emotions(
    class_id: str = Form(...),
    record_date: str = Form(...)
):
    try:
        supabase.table("emotion_votes").delete().eq("class_id", class_id).eq("record_date", record_date).execute()
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при нулиране на емоциите: {str(e)}")